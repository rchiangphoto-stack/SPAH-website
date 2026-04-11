import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backlink Opportunities"

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
normal = Font(name="Arial", size=10)
bold = Font(name="Arial", size=10, bold=True)
wrap = Alignment(vertical="top", wrap_text=True)
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

headers = ["Priority", "Platform", "Type", "Free/Paid", "SEO Value", "URL", "How to Get Listed", "Contact Info", "Status", "Notes"]
col_widths = [10, 26, 18, 12, 14, 42, 42, 32, 14, 30]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Priority color fills
fills = {
    "P1": PatternFill("solid", fgColor="FFCDD2"),
    "P2": PatternFill("solid", fgColor="FFF3E0"),
    "P3": PatternFill("solid", fgColor="FFFDE7"),
    "P4": PatternFill("solid", fgColor="E8F5E9"),
    "P5": PatternFill("solid", fgColor="E3F2FD"),
}

seo_fills = {
    "Very High": PatternFill("solid", fgColor="C8E6C9"),
    "High": PatternFill("solid", fgColor="DCEDC8"),
    "Moderate-High": PatternFill("solid", fgColor="F0F4C3"),
    "Moderate": PatternFill("solid", fgColor="FFF9C4"),
    "Low-Moderate": PatternFill("solid", fgColor="FFE0B2"),
}

listings = [
    # P1 - Do these immediately (free, high value)
    ["P1", "Yelp for Business", "Directory", "Free", "Very High",
     "https://business.yelp.com/",
     "Claim listing > complete all profile fields > upload photos > respond to reviews",
     "business.yelp.com", "Not Started", "DA 90+. Yelp pages rank on page 1 for local searches. MUST claim & optimize."],

    ["P1", "Patch.com (Alhambra)", "Local News/Directory", "Free", "High",
     "https://patch.com/california/los-angeles/directory",
     "Sign up > Get Posting > select Business > verify account > create listing",
     "patch.com/sign-up", "Not Started", "DA 80+. Business listings rank well in local Google searches."],

    ["P1", "Alignable", "Business Network", "Free", "Moderate",
     "https://www.alignable.com/",
     "Sign up free > complete business profile > join local groups > get recommendations",
     "alignable.com", "Not Started", "LinkedIn for local businesses. 10M+ business owners. Profiles indexed by Google."],

    ["P1", "AKC Veterinary Network", "Vet Directory", "Free", "Moderate-High",
     "https://www.apps.akc.org/vetnet/main",
     "Enroll at akc.org/vetnet > appear in AKC Vet Finder directory",
     "akc.org/vetnet", "Not Started", "DA 80+. Free enrollment. ~2,000 practices listed. Great authority link."],

    ["P1", "CareCredit Provider Directory", "Healthcare Directory", "Free", "Moderate-High",
     "https://www.carecredit.com/providers/animal-healthcare/",
     "Call 800-300-3046 (opt 5) or visit carecredit.com/vetenroll",
     "(800) 300-3046 opt 5", "Not Started", "You already have a CareCredit backlink. Verify listing has new Alhambra address."],

    ["P1", "BringFido", "Pet Travel Directory", "Free", "Moderate",
     "https://www.bringfido.com/resource/veterinarians/",
     "Use 'Post a Spot' feature > submit business > reviewed in 5-7 days",
     "bringfido.com", "Not Started", "Dofollow backlink. Good for visibility to traveling pet owners."],

    ["P1", "VetLocator.com", "Vet Directory", "Free", "Moderate",
     "https://www.vetlocator.com/",
     "Submit clinic through the site (60,000+ vets listed)",
     "vetlocator.com", "Not Started", "Free niche directory with relevant backlink."],

    ["P1", "Great Pet Care Vet Finder", "Vet Directory", "Free", "Moderate",
     "https://vets.greatpetcare.com/",
     "Search and claim your listing on the site",
     "vets.greatpetcare.com", "Not Started", "Free to claim. Niche pet care directory."],

    ["P1", "NewsBreak (Alhambra)", "Local News", "Free", "Moderate-High",
     "https://www.newsbreak.com/alhambra-ca",
     "Create free contributor account > publish press release about the move/grand opening",
     "newsbreak.com", "Not Started", "Local news aggregator. Free to publish. Great for move announcement."],

    # P2 - Chamber of Commerce (paid but high local SEO value)
    ["P2", "Alhambra Chamber of Commerce", "Chamber", "Paid (membership)", "High",
     "https://www.alhambrachamber.org/join-us",
     "Contact chamber > join as member > get directory listing + Around Alhambra feature",
     "hello@alhambrachamber.org | (626) 282-8481", "Not Started", "Strong local citation. .org domain. Also publishes Around Alhambra monthly."],

    ["P2", "South Pasadena Chamber", "Chamber", "Paid (membership)", "High",
     "https://southpasadena.net/",
     "Join as member > get directory listing",
     "info@southpasadena.net | (626) 441-2339", "Not Started", "Maintain connection to SP community even after move."],

    ["P2", "San Gabriel Chamber", "Chamber", "Paid (membership)", "High",
     "https://sangabrielchamber.org/",
     "Contact chamber > join > directory listing",
     "sangabrielchamber.org | 620 W Santa Anita Ave, San Gabriel", "Not Started", "Covers San Gabriel area you serve."],

    ["P2", "Pasadena Chamber of Commerce", "Chamber", "Paid (membership)", "High",
     "https://www.pasadena-chamber.org/forms/application-membership",
     "Submit membership application online",
     "pasadena-chamber.org", "Not Started", "Large, authoritative local chamber."],

    ["P2", "Regional Chamber SGV", "Chamber", "Paid (membership)", "High",
     "https://www.rccsgv.com",
     "Contact at (626) 810-8476 to join",
     "(626) 810-8476", "Not Started", "Covers entire San Gabriel Valley region."],

    # P3 - Requires certification or more effort
    ["P3", "AVMA Directory", "Professional Org", "Paid (membership)", "High",
     "https://www.avma.org/membership",
     "Become AVMA member > practice appears in Member Directory",
     "(800) 248-2862 | memberrecords@avma.org", "Not Started", "DA 70+. Requires AVMA membership dues. Check if already a member."],

    ["P3", "Fear Free Directory", "Certification Directory", "Paid (certification)", "Moderate",
     "https://directory.fearfree.com/",
     "Complete Fear Free certification (online modules, 80%+ score, sign pledge)",
     "certifyme@fearfree.com", "Not Started", "Requires full certification. Good marketing badge + directory listing."],

    # P4 - Lower priority / newer platforms
    ["P4", "Around Alhambra (feature article)", "Local Publication", "Free (via chamber)", "Moderate-High",
     "https://aroundalhambra.org/",
     "Pitch a business spotlight story about SPAH moving to Alhambra",
     "hello@alhambrachamber.org", "Not Started", "Published by Alhambra Chamber. Contextual backlink from local publication."],

    ["P4", "Downtown Alhambra", "Local Directory", "Free", "Moderate",
     "https://www.downtownalhambra.com/",
     "Contact for business feature / event listing opportunities",
     "downtownalhambra.com", "Not Started", "Local community site. Good for neighborhood visibility."],

    ["P4", "ZooDoc", "Booking Platform", "Free", "Low-Moderate",
     "https://zoodoc.org/",
     "Register as a provider on zoodoc.org",
     "zoodoc.org", "Not Started", "Growing platform. 'Zocdoc for pets.' Worth claiming early."],

    ["P4", "Petworks", "Booking Platform", "Free", "Low-Moderate",
     "https://www.petworks.com/",
     "Create provider profile on petworks.com",
     "petworks.com", "Not Started", "Browse/book appointments platform. Free profile."],

    ["P4", "Thumbtack", "Service Directory", "Free (pay-per-lead)", "Low-Moderate",
     "https://www.thumbtack.com/",
     "Create professional profile > get reviews for visibility",
     "thumbtack.com", "Not Started", "Not primary vet directory but high DA. Free to list."],
]

for row_idx, data in enumerate(listings, 2):
    priority = data[0]
    seo_val = data[4]
    p_fill = fills.get(priority, PatternFill())
    s_fill = seo_fills.get(seo_val, PatternFill())
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx in [1, 4, 5, 9]:
            cell.alignment = center_wrap
        if col_idx == 1:
            cell.fill = p_fill
            cell.font = bold
        if col_idx == 5:
            cell.fill = s_fill

# Status dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Submitted,Live,Skipped"', allow_blank=True)
ws.add_data_validation(dv)
for row in range(2, len(listings) + 2):
    dv.add(ws.cell(row=row, column=9))

# Summary
sr = len(listings) + 3
ws.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True, size=12, name="Arial")
last = len(listings) + 1

labels = ["Total Opportunities:", "Free:", "Paid:", "Submitted/Live:", "Not Started:"]
formulas = [
    f'=COUNTA(I2:I{last})',
    f'=COUNTIF(D2:D{last},"Free")',
    f'=COUNTIFS(D2:D{last},"Paid*")',
    f'=COUNTIF(I2:I{last},"Submitted")+COUNTIF(I2:I{last},"Live")',
    f'=COUNTIF(I2:I{last},"Not Started")',
]

for i, (label, formula) in enumerate(zip(labels, formulas)):
    r = sr + 1 + i
    ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=2, value=formula).font = normal

ws.auto_filter.ref = f"A1:J{last}"
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30
for row in range(2, len(listings) + 2):
    ws.row_dimensions[row].height = 60

# ── Sheet 2: Current Backlinks ──
ws2 = wb.create_sheet("Current Backlinks (50)")

headers2 = ["Site", "Linking Pages", "Target Pages", "Quality", "Action Needed"]
col_widths2 = [30, 16, 16, 16, 40]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin

for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

good_fill = PatternFill("solid", fgColor="C8E6C9")
neutral_fill = PatternFill("solid", fgColor="FFF9C4")
spam_fill = PatternFill("solid", fgColor="FFCDD2")

backlinks = [
    ["reddit.com", 26, 3, "Good", "Keep engaging in local subreddits (r/pasadena, r/alhambra, r/LosAngeles)"],
    ["guide.in.ua", 4, 1, "Spam", "Ignore - low quality foreign domain"],
    ["luke.lol", 3, 3, "Spam", "Ignore - low quality domain"],
    ["mapquest.com", 2, 1, "Good", "Update address to Alhambra location"],
    ["mycritters.com", 2, 2, "Good", "Niche pet directory - verify listing is current"],
    ["petinsurancereview.com", 2, 1, "Good", "Pet niche - verify listing info"],
    ["bunchiaburrow.com", 1, 1, "Neutral", "Small pet site - no action needed"],
    ["carecredit.com", 1, 1, "Good", "High DA - verify provider listing has new address"],
    ["catposse.org", 1, 1, "Good", "Cat rescue org - maintain relationship"],
    ["catvetfinder.com", 1, 1, "Good", "Vet directory - update address"],
    ["dogtrainermanhattan.com", 1, 1, "Neutral", "Out of area - no action needed"],
    ["getlocalverified.com", 1, 1, "Neutral", "Local directory - verify listing"],
    ["nextdoor.com", 1, 1, "Good", "Strong local signal - keep Nextdoor profile active"],
    ["poultrydvm.com", 1, 1, "Good", "Exotic/poultry vet directory - great niche link"],
    ["topvet.net", 1, 1, "Good", "Vet directory - verify listing"],
    ["veterinarioscercademi.net", 1, 1, "Spam", "Foreign domain - ignore"],
    ["z1biz.com", 1, 1, "Neutral", "Small business directory - no action needed"],
]

quality_fills = {"Good": good_fill, "Neutral": neutral_fill, "Spam": spam_fill}

for row_idx, data in enumerate(backlinks, 2):
    quality = data[3]
    for col_idx, val in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx in [2, 3, 4]:
            cell.alignment = center_wrap
        if col_idx == 4:
            cell.fill = quality_fills.get(quality, PatternFill())

sr2 = len(backlinks) + 3
ws2.cell(row=sr2, column=1, value="BACKLINK SUMMARY").font = Font(bold=True, size=12, name="Arial")
ws2.cell(row=sr2+1, column=1, value="Total External Links:").font = bold
ws2.cell(row=sr2+1, column=2, value=50).font = normal
ws2.cell(row=sr2+2, column=1, value="Good Quality:").font = bold
ws2.cell(row=sr2+2, column=2, value=f'=COUNTIF(D2:D{len(backlinks)+1},"Good")').font = normal
ws2.cell(row=sr2+3, column=1, value="Spam/Low Quality:").font = bold
ws2.cell(row=sr2+3, column=2, value=f'=COUNTIF(D2:D{len(backlinks)+1},"Spam")').font = normal
ws2.cell(row=sr2+4, column=1, value="Need Address Update:").font = bold
ws2.cell(row=sr2+4, column=2, value=4).font = normal
ws2.cell(row=sr2+4, column=3, value="mapquest, carecredit, catvetfinder, topvet").font = Font(name="Arial", size=9, italic=True)

ws2.freeze_panes = "A2"

output = "C:/Users/rchia/Documents/SPAH-website/SPAH-Backlink-Opportunities.xlsx"
wb.save(output)
print(f"Created: {output}")
