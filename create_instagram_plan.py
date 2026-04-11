import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Sheet 1: Content Calendar Overview ──
ws = wb.active
ws.title = "Content Calendar"

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
normal = Font(name="Arial", size=10)
bold = Font(name="Arial", size=10, bold=True)
wrap = Alignment(vertical="top", wrap_text=True)
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

headers = ["Day", "Date", "Theme", "Post Type", "Visual Idea", "Caption", "Hashtags", "Best Time", "Status"]
col_widths = [8, 14, 22, 14, 40, 60, 50, 12, 14]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

day_fills = [
    PatternFill("solid", fgColor="FFF8E1"),  # Mon - warm yellow
    PatternFill("solid", fgColor="E8F5E9"),  # Tue - green
    PatternFill("solid", fgColor="FFEBEE"),  # Wed - red (move day!)
    PatternFill("solid", fgColor="F3E5F5"),  # Thu - purple
    PatternFill("solid", fgColor="E3F2FD"),  # Fri - blue
    PatternFill("solid", fgColor="FFF3E0"),  # Sat - orange
    PatternFill("solid", fgColor="FCE4EC"),  # Sun - pink
]

posts = [
    {
        "day": "Mon",
        "date": "Mar 30",
        "theme": "Countdown / Teaser",
        "type": "Photo or Reel",
        "visual": "Photo of new Alhambra clinic with \"3 Days\" countdown overlay. Behind-the-scenes: boxes, equipment being placed, team prepping. Raw energy.",
        "caption": "3 days. \U0001f3e5\n\nAfter 30+ years on Fremont Ave, we're bringing South Pasadena Animal Hospital to the heart of Alhambra.\n\nNew address. Same doctors. Same compassionate care for pets of ALL shapes and sizes \u2014 yes, even your chinchilla.\n\n\U0001f4cd 3116 W Main St, Alhambra\n\U0001f5d3\ufe0f Opening April 1st\n\nDrop a \U0001f43e if you're coming with us.",
        "hashtags": "#SPAHisMoving #AlhambraVet #SouthPasadenaAnimalHospital #NewLocation #VetClinic #AlhambraCA #ExoticVet #PetCare #SGV #SanGabrielValley #VetLife #AnimalHospital #WeMovedVet #PetLovers #AlhambraAnimalHospital",
        "time": "6\u20139 PM",
    },
    {
        "day": "Tue",
        "date": "Mar 31",
        "theme": "Meet the Doctors",
        "type": "Carousel",
        "visual": "Swipeable carousel \u2014 each slide is one doctor with photo + name + title + one fun fact. Consistent template with brand colors.",
        "caption": "The faces behind your pet's care aren't changing \u2014 just the zip code. \U0001f44b\n\nSwipe to meet the team moving to our brand new Alhambra clinic tomorrow.\n\nWhether your best friend has four legs, scales, feathers, or a shell \u2014 we've got a doctor who knows them inside and out.\n\nFun fact: our team has collectively treated over 50 different species. What's the most unusual pet you've brought to us? Tell us below \U0001f447",
        "hashtags": "#MeetTheVet #ExoticAnimalVet #AlhambraVet #VetTeam #SouthPasadenaAnimalHospital #ReptileVet #BirdVet #SmallAnimalVet #PasadenaVet #SGVPets #VeterinaryCare #AnimalDoctor #PetHealth #AlhambraCA #CompanionAnimalCare",
        "time": "6\u20139 PM",
    },
    {
        "day": "Wed",
        "date": "Apr 1",
        "theme": "MOVE DAY!",
        "type": "Reel + Stories",
        "visual": "Reel/photo dump: team unlocking doors, ribbon cut moment, first patient walking in, before/after old vs new clinic. High energy, celebratory. Go LIVE for 5\u201310 min.",
        "caption": "IT'S OFFICIAL. \U0001f389\U0001f511\n\nSouth Pasadena Animal Hospital is now open at our brand new Alhambra location.\n\nSame trusted care since 1990. Bigger space. Better equipped. Easier to reach from Alhambra, Monterey Park, San Gabriel, South Pasadena, and beyond.\n\n\U0001f4cd 3116 W Main St, Alhambra, CA 91801\n\U0001f4de (626) 799-1123\n\U0001f4bb Book online \u2014 link in bio\n\nWe're here. Bring your babies. \U0001f436\U0001f431\U0001f430\U0001f98e\n\n#NowOpen",
        "hashtags": "#GrandOpening #NowOpen #AlhambraVet #SPAH #SouthPasadenaAnimalHospital #AlhambraAnimalHospital #VetClinic #NewClinic #PetCareAlhambra #MontereyParkVet #SanGabrielVet #SGV #SouthPasadena #VetLife #SupportLocal #SmallBusiness #PetHealth",
        "time": "12 PM + 6 PM",
    },
    {
        "day": "Thu",
        "date": "Apr 2",
        "theme": "Exotic Pet Spotlight",
        "type": "Photo",
        "visual": "Stunning photo of an exotic patient (bearded dragon, rabbit, chinchilla, bird) with owner permission. Or high-quality image with text overlay: \"Yes, we treat them too.\"",
        "caption": "Dogs and cats aren't the only ones who deserve great healthcare.\n\nAt SPAH, we're one of the few clinics in the San Gabriel Valley that treats exotic pets \u2014 rabbits, reptiles, birds, guinea pigs, hamsters, and more.\n\nHard to find a vet for your unusual buddy? Not anymore.\n\n\U0001f4cd Now at 3116 W Main St, Alhambra\n\U0001f517 Book an exotic pet wellness exam \u2014 link in bio",
        "hashtags": "#ExoticVet #ExoticPets #ReptileVet #BirdVet #RabbitVet #BeardedDragon #GuineaPig #Chinchilla #ExoticAnimalCare #AlhambraVet #PasadenaExoticVet #SGVExoticVet #UnusualPets #PocketPets #SmallAnimalVet #ExoticPetCare #HerpVet",
        "time": "6\u20139 PM",
    },
    {
        "day": "Fri",
        "date": "Apr 3",
        "theme": "Neighborhood / Community",
        "type": "Graphic or Photo",
        "visual": "Map graphic with pin on Alhambra + radiating lines to surrounding cities. Or photo of Main St / local Alhambra landmarks near the clinic. Vibe: \"We're your neighbor now.\"",
        "caption": "New to the neighborhood, but not new to the community. \U0001f3d8\ufe0f\n\nWe've been caring for pets across the San Gabriel Valley for over 35 years. Now we're even closer to:\n\n\U0001f539 Alhambra \u2014 2 min\n\U0001f539 Monterey Park \u2014 5 min\n\U0001f539 San Gabriel \u2014 5 min\n\U0001f539 South Pasadena \u2014 7 min\n\U0001f539 Rosemead \u2014 8 min\n\U0001f539 San Marino \u2014 7 min\n\U0001f539 Highland Park \u2014 8 min\n\U0001f539 Eagle Rock \u2014 10 min\n\nNo matter which city you call home, quality vet care just got closer.\n\n\U0001f4cd 3116 W Main St, Alhambra",
        "hashtags": "#AlhambraCA #MontereyPark #SanGabriel #SouthPasadena #Rosemead #SanMarino #HighlandPark #EagleRock #SGV #SanGabrielValley #LocalVet #CommunityVet #NearMe #VetNearMe #AlhambraAnimalHospital #SupportLocalBusiness #PetCommunity",
        "time": "6\u20139 PM",
    },
    {
        "day": "Sat",
        "date": "Apr 4",
        "theme": "Services Showcase",
        "type": "Carousel",
        "visual": "Carousel with each slide = one service with clean icon + short description. Services: Wellness, Vaccinations, Surgery, Dental, Dermatology, Exotic, Diagnostics. Brand colors (navy + warm).",
        "caption": "Everything your pet needs \u2014 under one roof. \U0001fa7a\n\n\u2705 Wellness & preventive exams\n\u2705 Vaccinations\n\u2705 Surgery (soft tissue & orthopedic)\n\u2705 Dental cleanings & extractions\n\u2705 Dermatology\n\u2705 In-house diagnostics (lab, X-ray, ultrasound)\n\u2705 Exotic animal medicine\n\nNew clients welcome. We'd love to meet your pet.\n\n\U0001f4cd 3116 W Main St, Alhambra\n\U0001f4de (626) 799-1123\n\U0001f517 Book online \u2014 link in bio",
        "hashtags": "#VetServices #PetDentist #DogDentist #PetSurgery #VetDermatology #PetWellness #Vaccinations #VetCare #AlhambraVet #FullServiceVet #AnimalHospital #PetHealth #PreventiveCare #VetDiagnostics #PetCheckup #HealthyPets #AlhambraCA",
        "time": "10 AM\u201312 PM",
    },
    {
        "day": "Sun",
        "date": "Apr 5",
        "theme": "Throwback / Gratitude",
        "type": "Carousel or Reel",
        "visual": "Side-by-side or carousel: old Fremont Ave clinic photos (building, memories, team throwbacks) next to new Alhambra clinic. Sentimental. Last slide: \"Book your first visit at our new home.\"",
        "caption": "One last look back before we look ahead. \u2764\ufe0f\n\nThank you, South Pasadena, for over three decades of trusting us with your fur babies (and scaly babies, and feathered babies).\n\nFremont Ave will always be where our story started. But Alhambra is where the next chapter begins \u2014 and we're bringing everything that made SPAH special with us.\n\nTo our longtime families: we're not going far.\nTo our new neighbors: we can't wait to meet you.\n\n\U0001f4cd 3116 W Main St, Alhambra, CA 91801\n\U0001f517 Book your pet's first visit at our new home \u2014 link in bio",
        "hashtags": "#ThrowbackThursday #VetLife #GratefulVet #SouthPasadena #AlhambraVet #SPAH #SouthPasadenaAnimalHospital #NewBeginnings #PetFamily #VetCommunity #SmallBusinessMove #SGV #ThankYou #PetLove #VetAppreciation #AnimalHospital #AlhambraCA",
        "time": "6\u20139 PM",
    },
]

# Fix the move day theme (emoji encoding issue)
posts[2]["theme"] = "MOVE DAY!"

for row_idx, p in enumerate(posts, 2):
    fill = day_fills[row_idx - 2]
    values = [p["day"], p["date"], p["theme"], p["type"], p["visual"], p["caption"], p["hashtags"], p["time"], "Not Posted"]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx <= 4 or col_idx == 8:
            cell.alignment = center_wrap
        if col_idx == 1 or col_idx == 3:
            cell.fill = fill
            cell.font = bold

# Move day row highlight
move_fill = PatternFill("solid", fgColor="FFCDD2")
for col in range(1, 10):
    ws.cell(row=4, column=col).fill = move_fill
ws.cell(row=4, column=1).fill = move_fill
ws.cell(row=4, column=3).fill = move_fill

# Status dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Posted,Scheduled,Posted"', allow_blank=True)
ws.add_data_validation(dv)
for row in range(2, 9):
    dv.add(ws.cell(row=row, column=9))

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30
for row in range(2, 9):
    ws.row_dimensions[row].height = 160

# ── Sheet 2: Captions (copy-paste ready) ──
ws2 = wb.create_sheet("Captions - Copy Paste")

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 80
ws2.column_dimensions["C"].width = 60

cap_headers = ["Day / Date", "Caption (copy & paste)", "Hashtags (copy & paste)"]
for col, h in enumerate(cap_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

for row_idx, p in enumerate(posts, 2):
    label = f'{p["day"]} {p["date"]}\n{p["theme"]}'
    ws2.cell(row=row_idx, column=1, value=label).font = bold
    ws2.cell(row=row_idx, column=1).alignment = center_wrap
    ws2.cell(row=row_idx, column=1).border = thin
    ws2.cell(row=row_idx, column=1).fill = day_fills[row_idx - 2]

    ws2.cell(row=row_idx, column=2, value=p["caption"]).font = normal
    ws2.cell(row=row_idx, column=2).alignment = wrap
    ws2.cell(row=row_idx, column=2).border = thin

    ws2.cell(row=row_idx, column=3, value=p["hashtags"]).font = normal
    ws2.cell(row=row_idx, column=3).alignment = wrap
    ws2.cell(row=row_idx, column=3).border = thin

    ws2.row_dimensions[row_idx].height = 200

ws2.freeze_panes = "A2"

# ── Sheet 3: Tips ──
ws3 = wb.create_sheet("Posting Tips")
ws3.column_dimensions["A"].width = 70

tips = [
    "INSTAGRAM POSTING TIPS FOR SPAH",
    "",
    "TIMING",
    "\u2022 Post between 6\u20139 PM for peak engagement (local/pet content)",
    "\u2022 Move Day (Apr 1): Post at 12 PM AND 6 PM for max reach",
    "\u2022 Saturday: Try 10 AM\u201312 PM (weekend browsing hours)",
    "",
    "FORMAT",
    "\u2022 Use Reels for Days 1, 3, and 7 \u2014 video gets 2x reach over static",
    "\u2022 Carousels (Days 2, 6) get highest saves and shares",
    "\u2022 Story every post with a poll or question sticker to boost engagement",
    "",
    "MOVE DAY SPECIAL (Apr 1)",
    "\u2022 Consider going LIVE for 5\u201310 min during the first patient visit",
    "\u2022 Post to Stories throughout the day (behind-the-scenes moments)",
    "\u2022 Use countdown sticker in Stories leading up to it",
    "",
    "LOCATION & TAGS",
    "\u2022 Tag location as \"Alhambra, California\" on every post starting Apr 1",
    "\u2022 Tag @southpasah on all posts",
    "\u2022 Tag pet owners in photos (with permission) for extra reach",
    "",
    "HASHTAG STRATEGY",
    "\u2022 Instagram allows up to 30 hashtags \u2014 use 15\u201320 per post",
    "\u2022 Mix broad (#VetLife, #PetCare) with local (#AlhambraVet, #SGV)",
    "\u2022 Put hashtags in the caption, not a separate comment (better for SEO)",
    "\u2022 Always include #AlhambraVet and #SouthPasadenaAnimalHospital",
    "",
    "ENGAGEMENT",
    "\u2022 Reply to every comment within 1 hour of posting",
    "\u2022 Like and comment on local Alhambra business pages",
    "\u2022 Share user-generated content (pet owners posting about visits)",
    "\u2022 Use question stickers: \"What's your pet's name?\" drives replies",
]

for i, tip in enumerate(tips, 1):
    cell = ws3.cell(row=i, column=1, value=tip)
    if tip in ["INSTAGRAM POSTING TIPS FOR SPAH"]:
        cell.font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    elif tip in ["TIMING", "FORMAT", "MOVE DAY SPECIAL (Apr 1)", "LOCATION & TAGS", "HASHTAG STRATEGY", "ENGAGEMENT"]:
        cell.font = Font(bold=True, size=11, name="Arial", color="1F4E79")
        cell.fill = PatternFill("solid", fgColor="E3F2FD")
    else:
        cell.font = normal

output = "C:/Users/rchia/Documents/SPAH-website/SPAH-Instagram-7Day-Plan.xlsx"
wb.save(output)
print(f"Created: {output}")
