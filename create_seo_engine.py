import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
normal = Font(name="Arial", size=10)
bold = Font(name="Arial", size=10, bold=True)
bold12 = Font(name="Arial", size=12, bold=True, color="1F4E79")
wrap = Alignment(vertical="top", wrap_text=True)
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

def style_headers(ws, headers, widths):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

# ═══════════════════════════════════════════════════
# SHEET 1: COMPETITIVE RESEARCH SUMMARY
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Competitive Research"

headers1 = ["Competitor", "Location", "Treats Exotics?", "Has Blog?", "Review Count", "Rating", "Threat Level", "Key Weakness", "SPAH Opportunity"]
widths1 = [28, 18, 14, 12, 14, 10, 14, 35, 40]
style_headers(ws1, headers1, widths1)

threat_fills = {
    "HIGH": PatternFill("solid", fgColor="FFCDD2"),
    "MEDIUM": PatternFill("solid", fgColor="FFF3E0"),
    "LOW": PatternFill("solid", fgColor="E8F5E9"),
}

competitors = [
    ["Exotic Animal Vet Center", "Pasadena", "YES (exotics only)", "YES (thin)", "~284 Yelp", "4+", "HIGH", "Exotics-ONLY (no dogs/cats), thin content", "SPAH treats ALL pets. Produce deeper exotic content."],
    ["Vet Villa Animal Hospital", "South Pasadena", "NO", "NO", "~450 Yelp", "4.5", "HIGH", "No exotics, no blog, no SEO content", "SPAH handles exotics. Start a blog to dominate search."],
    ["Veterinary Healthcare Center", "Monterey Park", "Some", "NO", "~690 Birdeye", "4.5", "MEDIUM", "No blog despite huge review count", "Create content targeting Monterey Park keywords."],
    ["Alhambra Vet Hospital", "Alhambra", "Some reptiles", "NO", "~208 Google", "4.5", "MEDIUM", "No blog, minimal exotic pages", "Target 'Alhambra animal hospital' with better content."],
    ["Monterey Park Animal Hospital", "Monterey Park", "YES", "Species pages only", "~114 Yelp", "4+", "MEDIUM", "Basic species pages, no deep content", "Write better exotic guides than their thin pages."],
    ["San Gabriel Animal Hospital", "San Gabriel", "Some mentions", "NO", "~449 Yelp", "4+", "MEDIUM", "Claims exotics but dog/cat focused", "Genuine exotic expertise is SPAH's advantage."],
    ["Petsadena Animal Hospital", "Pasadena", "NO", "NO", "~385 Yelp", "4.5", "LOW", "Dogs/cats only, no content", "Different niche - not direct competition on exotics."],
    ["VCA Mission", "Alhambra", "NO", "Corporate (generic)", "~284 Yelp", "4", "LOW", "Corporate chain, no local content, no exotics", "SPAH is independent + treats exotics. Highlight both."],
    ["West Main Animal Hospital", "Alhambra", "NO (refers out)", "NO", "~139 Yelp", "4", "LOW", "Refers exotics to other clinics", "SPAH IS the clinic they refer exotics to."],
    ["Eagle Rock Pet Care", "Eagle Rock", "YES", "Unknown", "Limited", "--", "LOW", "Small practice, limited web presence", "Target Eagle Rock geo keywords."],
]

for row_idx, data in enumerate(competitors, 2):
    threat = data[6]
    for col_idx, val in enumerate(data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx in [3, 4, 5, 6, 7]:
            cell.alignment = center_wrap
        if col_idx == 7:
            cell.fill = threat_fills.get(threat, PatternFill())
            cell.font = bold

# Key findings
r = len(competitors) + 3
findings = [
    "KEY FINDINGS",
    "",
    "1. ALMOST NO COMPETITORS HAVE BLOGS. This is the #1 opportunity.",
    "2. No local vet has comprehensive exotic pet content. SPAH can own this space.",
    "3. SPAH does NOT rank for 'exotic vet Alhambra' or 'reptile vet San Gabriel Valley' yet.",
    "4. Exotic Animal Vet Center (Pasadena) is the only real exotic content competitor - but their content is thin.",
    "5. SPAH's review count (~82 Google) is well behind Vet Villa (~450) and VHC (~690). Need to build reviews.",
    "6. SPAH has a unique advantage: treats BOTH regular pets AND exotics (one-stop shop for multi-pet families).",
    "7. The 'reptile vet Los Angeles' and 'bird vet Pasadena' search landscape has zero local blog content.",
    "8. Content targeting symptom-based searches ('bearded dragon not eating', 'guinea pig sneezing') has massive volume and zero local competition.",
]
for i, f in enumerate(findings):
    cell = ws1.cell(row=r+i, column=1, value=f)
    if i == 0:
        cell.font = bold12
    elif f.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.")):
        cell.font = bold
    else:
        cell.font = normal
    ws1.merge_cells(start_row=r+i, start_column=1, end_row=r+i, end_column=9)

for row in range(2, len(competitors) + 2):
    ws1.row_dimensions[row].height = 45

# ═══════════════════════════════════════════════════
# SHEET 2: BLOG TITLES & KEYWORDS (25 posts)
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Blog Titles & Keywords")

headers2 = ["#", "Blog Title", "Primary Keyword", "Supporting Keywords", "Search Intent", "Difficulty", "Priority", "Target Species", "Word Count", "Status"]
widths2 = [5, 50, 28, 50, 16, 12, 10, 16, 12, 14]
style_headers(ws2, headers2, widths2)

intent_fills = {
    "Informational": PatternFill("solid", fgColor="E3F2FD"),
    "Commercial": PatternFill("solid", fgColor="F3E5F5"),
    "BOFU": PatternFill("solid", fgColor="FFCDD2"),
    "Local": PatternFill("solid", fgColor="E8F5E9"),
}
diff_fills = {
    "Low": PatternFill("solid", fgColor="C8E6C9"),
    "Medium": PatternFill("solid", fgColor="FFF9C4"),
    "High": PatternFill("solid", fgColor="FFCDD2"),
}
pri_fills = {
    "P1": PatternFill("solid", fgColor="FFCDD2"),
    "P2": PatternFill("solid", fgColor="FFF3E0"),
    "P3": PatternFill("solid", fgColor="E8F5E9"),
}

blogs = [
    # P1 - Publish first (highest impact, lowest competition)
    [1, "Signs Your Bearded Dragon Needs to See a Vet", "bearded dragon vet", "bearded dragon sick signs, bearded dragon not eating, reptile vet Alhambra, bearded dragon health", "Informational", "Low", "P1", "Reptile", "1500-2000", "Not Started"],
    [2, "Guinea Pig Sneezing: When to Worry and When It's Normal", "guinea pig sneezing", "guinea pig upper respiratory infection, guinea pig vet, exotic vet near me, guinea pig sick symptoms", "Informational", "Low", "P1", "Small Mammal", "1200-1500", "Not Started"],
    [3, "Rabbit GI Stasis: Symptoms, Treatment & Prevention", "rabbit GI stasis", "rabbit not eating, rabbit stomach problems, rabbit vet Alhambra, rabbit emergency signs", "Informational", "Low", "P1", "Rabbit", "1500-2000", "Not Started"],
    [4, "Why Is My Bird Pulling Out Its Feathers?", "bird feather plucking", "parrot plucking feathers, bird vet near me, avian vet Pasadena, bird self-mutilation", "Informational", "Low", "P1", "Bird", "1200-1500", "Not Started"],
    [5, "What to Expect at Your Exotic Pet's First Vet Visit", "exotic pet vet visit", "first vet visit exotic pet, exotic vet Alhambra, exotic pet wellness exam cost, new exotic pet owner", "Commercial", "Low", "P1", "All Exotics", "1200-1500", "Not Started"],

    # P1 continued - species care guides
    [6, "The Complete Guide to Bearded Dragon Care in Southern California", "bearded dragon care guide", "bearded dragon husbandry, bearded dragon diet, bearded dragon lighting, reptile care SoCal", "Informational", "Low", "P1", "Reptile", "2000-2500", "Not Started"],
    [7, "5 Foods That Are Toxic to Pet Birds (and What to Feed Instead)", "foods toxic to birds", "can birds eat avocado, bird diet, safe foods for parrots, avian nutrition", "Informational", "Low", "P1", "Bird", "1200-1500", "Not Started"],
    [8, "How to Tell If Your Snake Is Sick: A Vet's Checklist", "sick snake symptoms", "snake vet near me, ball python not eating, snake respiratory infection, reptile vet Alhambra", "Informational", "Low", "P1", "Reptile", "1500-1800", "Not Started"],

    # P2 - High value, build authority
    [9, "Do Guinea Pigs Need Annual Vet Checkups?", "guinea pig vet checkup", "guinea pig wellness exam, how often vet guinea pig, exotic pet annual exam cost", "Commercial", "Low", "P2", "Small Mammal", "1000-1200", "Not Started"],
    [10, "Rabbit Hemorrhagic Disease (RHDV2): What LA Pet Owners Need to Know", "RHDV2 vaccine rabbit", "rabbit hemorrhagic disease Los Angeles, RHDV vaccine near me, rabbit vaccination schedule", "Informational", "Low", "P2", "Rabbit", "1500-1800", "Not Started"],
    [11, "Bearded Dragon Brumation vs. Sick: How to Tell the Difference", "bearded dragon brumation", "bearded dragon sleeping a lot, bearded dragon lethargic, is my bearded dragon hibernating", "Informational", "Low", "P2", "Reptile", "1200-1500", "Not Started"],
    [12, "How Much Does an Exotic Vet Visit Cost? A Transparent Guide", "exotic vet visit cost", "exotic pet vet prices, reptile vet cost, bird vet cost, affordable exotic vet", "Commercial", "Low", "P2", "All Exotics", "1200-1500", "Not Started"],
    [13, "Tortoise Shell Rot: Causes, Treatment & When to See a Vet", "tortoise shell rot", "turtle shell problems, tortoise vet, shell infection reptile, reptile vet Alhambra", "Informational", "Medium", "P2", "Reptile", "1200-1500", "Not Started"],
    [14, "The Best Diet for Pet Rabbits: A Vet's Complete Guide", "rabbit diet guide", "what to feed pet rabbit, rabbit nutrition, hay for rabbits, rabbit food list", "Informational", "Medium", "P2", "Rabbit", "1500-2000", "Not Started"],
    [15, "Can Guinea Pigs Eat [X]? The Ultimate Safe Food List", "guinea pig safe foods", "can guinea pigs eat strawberries, guinea pig diet, guinea pig vegetables, guinea pig fruit", "Informational", "Medium", "P2", "Small Mammal", "1500-2000", "Not Started"],
    [16, "Why Your Cat Keeps Scratching: Dermatology Guide for Pet Owners", "cat scratching dermatology", "cat skin allergies, cat itching, cat dermatitis, cat vet Alhambra dermatology", "Informational", "Medium", "P2", "Cat", "1200-1500", "Not Started"],

    # P2 - Local SEO content
    [17, "Finding an Exotic Vet in the San Gabriel Valley: What to Look For", "exotic vet San Gabriel Valley", "exotic vet near me SGV, reptile vet Pasadena, bird vet Alhambra, exotic animal hospital", "Local", "Low", "P2", "All Exotics", "1200-1500", "Not Started"],
    [18, "Why Alhambra Pet Owners Are Choosing a Non-Corporate Vet", "independent vet Alhambra", "non-corporate vet, private vet practice, Alhambra animal hospital, independent veterinarian", "Local", "Low", "P2", "All Pets", "1000-1200", "Not Started"],

    # P3 - Seasonal & topical
    [19, "Summer Heat Safety: Keeping Your Exotic Pets Cool in SoCal", "exotic pets heat safety", "keep reptile cool summer, guinea pig heat stroke, rabbit hot weather, SoCal pet care summer", "Informational", "Low", "P3", "All Exotics", "1200-1500", "Not Started"],
    [20, "Holiday Foods That Are Dangerous for Pets (Dogs, Cats & Exotics)", "holiday foods toxic pets", "Thanksgiving pet safety, Christmas foods bad for dogs, holiday pet dangers", "Informational", "Medium", "P3", "All Pets", "1500-1800", "Not Started"],
    [21, "Backyard Chickens in Los Angeles: Health Care & Vet Visits", "backyard chicken vet LA", "chicken vet Los Angeles, poultry vet near me, backyard flock health, chicken wellness exam", "Commercial", "Low", "P3", "Poultry", "1200-1500", "Not Started"],
    [22, "Dog Dental Disease: Why Your Dog's Bad Breath Might Be Serious", "dog dental disease", "dog bad breath, dog teeth cleaning cost, dog dental vet, periodontal disease dogs", "Informational", "Medium", "P3", "Dog", "1200-1500", "Not Started"],
    [23, "Hamster Health 101: Common Issues and When to See a Vet", "hamster vet", "hamster sick signs, hamster health problems, hamster wet tail, small mammal vet near me", "Informational", "Low", "P3", "Small Mammal", "1200-1500", "Not Started"],
    [24, "Why Regular Wellness Exams Save You Money (and Your Pet's Life)", "pet wellness exam benefits", "annual vet checkup important, preventive vet care, wellness exam cost, regular vet visits", "Commercial", "Medium", "P3", "All Pets", "1000-1200", "Not Started"],
    [25, "First-Time Reptile Owner? 7 Things Your Vet Wants You to Know", "first time reptile owner", "new reptile owner tips, beginner reptile care, reptile husbandry basics, reptile vet advice", "Informational", "Low", "P3", "Reptile", "1200-1500", "Not Started"],
]

for row_idx, data in enumerate(blogs, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx in [1, 5, 6, 7, 8, 9, 10]:
            cell.alignment = center_wrap
        if col_idx == 5:
            cell.fill = intent_fills.get(val, PatternFill())
        if col_idx == 6:
            cell.fill = diff_fills.get(val, PatternFill())
        if col_idx == 7:
            cell.fill = pri_fills.get(val, PatternFill())
            cell.font = bold
    ws2.row_dimensions[row_idx].height = 50

# Status dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv2 = DataValidation(type="list", formula1='"Not Started,Drafting,In Review,Published,Scheduled"', allow_blank=True)
ws2.add_data_validation(dv2)
for row in range(2, len(blogs) + 2):
    dv2.add(ws2.cell(row=row, column=10))

# ═══════════════════════════════════════════════════
# SHEET 3: CONTENT GENERATION PLAN (first 8 posts detailed)
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Content Plan (First 8)")

headers3 = ["Blog #", "Title", "H1", "H2 Sections", "H3 Subsections", "Internal Links", "CTA Placement", "Key Facts to Include"]
widths3 = [8, 40, 40, 55, 55, 35, 30, 50]
style_headers(ws3, headers3, widths3)

content_plans = [
    [1, "Signs Your Bearded Dragon Needs to See a Vet",
     "Signs Your Bearded Dragon Needs to See a Vet: A Veterinarian's Guide",
     "1. Normal vs Abnormal Behavior\n2. 8 Warning Signs That Need a Vet Visit\n3. Brumation vs Illness\n4. When It's an Emergency\n5. What Happens at a Bearded Dragon Vet Visit\n6. How to Find a Reptile Vet Near You",
     "Under Warning Signs: Not eating, Lethargy, Black beard, Mouth gaping, Eye issues, Skin discoloration, Weight loss, Abnormal stool\nUnder Emergency: Prolapse, Egg binding, Severe MBD, Trauma",
     "services.html (Exotic Pet Care)\nvet-alhambra.html\npricing.html\nBlog #6 (bearded dragon care guide)\nBlog #11 (brumation vs sick)",
     "After Section 2: 'Worried about your bearded dragon? Book a reptile wellness exam.'\nEnd of post: Full CTA block with phone + booking link",
     "Dr. Navia and Dr. Eng specialize in exotic pets. SPAH sees bearded dragons regularly. Include real symptoms vets actually see. Mention Southern California climate factors."],

    [2, "Guinea Pig Sneezing: When to Worry and When It's Normal",
     "Guinea Pig Sneezing: When to Worry and When It's Normal",
     "1. Why Guinea Pigs Sneeze (Normal Reasons)\n2. Upper Respiratory Infections in Guinea Pigs\n3. Signs It's More Than Just a Sneeze\n4. Diagnosis & Treatment\n5. Prevention Tips\n6. When to Call an Exotic Vet",
     "Under Normal: Dust, bedding, new scent, clearing nose\nUnder URI: Symptoms list, bacterial vs viral, Bordetella\nUnder Prevention: Cage hygiene, bedding type, vitamin C, temperature",
     "services.html (Exotic Pet Care)\nvet-alhambra.html\npricing.html\nBlog #9 (guinea pig checkups)\nBlog #15 (guinea pig foods)",
     "After Section 3: 'Is your guinea pig showing these symptoms? Schedule an exotic pet exam.'\nEnd: Full CTA",
     "Guinea pig URIs can be fatal if untreated. Mention Bordetella and Streptococcus. Vitamin C deficiency connection. SPAH treats guinea pigs regularly."],

    [3, "Rabbit GI Stasis: Symptoms, Treatment & Prevention",
     "Rabbit GI Stasis: A Vet's Guide to Symptoms, Treatment & Prevention",
     "1. What Is GI Stasis?\n2. Causes of GI Stasis\n3. Symptoms to Watch For\n4. Why GI Stasis Is an Emergency\n5. How Vets Treat GI Stasis\n6. Prevention: Diet & Lifestyle\n7. RHDV2 and Other Rabbit Health Concerns",
     "Under Causes: Low fiber, stress, pain, dehydration, dental disease\nUnder Symptoms: Not eating, small/no droppings, hunched posture, grinding teeth\nUnder Treatment: Fluid therapy, motility drugs, pain management, syringe feeding",
     "services.html\nvet-alhambra.html\nBlog #10 (RHDV2)\nBlog #14 (rabbit diet)\npricing.html",
     "After Section 4: 'Rabbit not eating? This could be an emergency. Call us at (626) 441-1314.'\nEnd: Full CTA",
     "GI stasis can kill within 24-48 hours. Emphasize urgency. Hay should be 80%+ of diet. SPAH has experience with rabbit emergencies."],

    [4, "Why Is My Bird Pulling Out Its Feathers?",
     "Why Is My Bird Pulling Out Its Feathers? A Veterinarian's Guide to Feather Plucking",
     "1. Understanding Feather Plucking\n2. Medical Causes\n3. Behavioral Causes\n4. Species Most Affected\n5. Diagnosis Process\n6. Treatment Options\n7. Prevention & Enrichment",
     "Under Medical: Skin infections, parasites, allergies, liver disease, nutritional deficiency\nUnder Behavioral: Boredom, stress, hormonal, sleep deprivation\nUnder Species: African greys, cockatoos, macaws",
     "services.html\nvet-alhambra.html\nBlog #7 (toxic bird foods)\npricing.html",
     "After Section 5: 'A vet exam is the first step. Book an avian wellness check.'\nEnd: Full CTA",
     "Dr. Navia has avian expertise. Feather destructive behavior is one of the most common avian vet complaints. Include that SPAH does beak trimming too."],

    [5, "What to Expect at Your Exotic Pet's First Vet Visit",
     "What to Expect at Your Exotic Pet's First Vet Visit",
     "1. Why Exotic Pets Need Vet Care\n2. How to Choose an Exotic Vet\n3. What to Bring to the Appointment\n4. What the Vet Will Do (By Species)\n5. Common Tests & Their Purpose\n6. How Often Should Exotic Pets See a Vet?\n7. Cost of an Exotic Pet Vet Visit",
     "Under By Species: Reptiles, birds, rabbits, guinea pigs, hamsters\nUnder Tests: Fecal, bloodwork, radiographs\nUnder Cost: Transparent pricing breakdown",
     "services.html\npricing.html (IMPORTANT - link directly)\nvet-alhambra.html\nabout.html (meet the doctors)",
     "After Section 2: 'SPAH treats birds, reptiles, rabbits, guinea pigs & more. Meet our exotic vet team.'\nEnd: Full CTA with 'New exotic patient? Book your first visit.'",
     "SPAH accepts new exotic patients. Pricing page has transparent costs. Dr. Navia specializes in exotics. Dr. Eng has 20+ years zoo medicine."],

    [6, "The Complete Guide to Bearded Dragon Care in Southern California",
     "The Complete Guide to Bearded Dragon Care in Southern California",
     "1. Why Bearded Dragons Thrive in SoCal\n2. Enclosure Setup\n3. Lighting & Temperature\n4. Diet & Nutrition\n5. Common Health Issues\n6. Finding a Reptile Vet in the SGV\n7. Annual Wellness Exam Checklist",
     "Under Enclosure: Tank size, substrate, decor\nUnder Lighting: UVB specifics, basking temps, SoCal sunlight considerations\nUnder Diet: Insects, greens, calcium, age-specific\nUnder Health: MBD, parasites, impaction, mouth rot",
     "services.html\nvet-alhambra.html\nBlog #1 (sick bearded dragon signs)\nBlog #11 (brumation)\npricing.html",
     "After Section 5: 'Concerned about your beardie? Schedule a reptile checkup.'\nSection 6: Natural CTA about SPAH's location\nEnd: Full CTA",
     "SoCal-specific tips (natural sunlight benefits, heat wave precautions). This is the cornerstone exotic content piece. Should be the most comprehensive beardie guide from any local vet."],

    [7, "5 Foods That Are Toxic to Pet Birds (and What to Feed Instead)",
     "5 Foods That Are Toxic to Pet Birds (and What to Feed Instead)",
     "1. Avocado\n2. Chocolate\n3. Caffeine\n4. Onions & Garlic\n5. Fruit Seeds & Pits\n6. Safe Foods Your Bird Will Love\n7. Signs of Poisoning: When to Call a Vet",
     "Under each toxic food: Why it's dangerous, symptoms of ingestion, how much is harmful\nUnder Safe Foods: Fruits, vegetables, pellets, treats by species\nUnder Poisoning Signs: Vomiting, lethargy, seizures, difficulty breathing",
     "services.html\nvet-alhambra.html\nBlog #4 (feather plucking)\npricing.html",
     "After Section 5: 'Think your bird ate something toxic? Call us immediately at (626) 441-1314.'\nEnd: Full CTA",
     "Include specific toxin names (persin in avocado, theobromine in chocolate). Practical and shareable content. High search volume for 'can birds eat' queries."],

    [8, "How to Tell If Your Snake Is Sick: A Vet's Checklist",
     "How to Tell If Your Snake Is Sick: A Veterinarian's Checklist",
     "1. Normal Snake Behavior\n2. 10 Signs Your Snake May Be Sick\n3. Common Snake Illnesses\n4. When Shedding Goes Wrong\n5. Respiratory Infections in Snakes\n6. What to Expect at a Snake Vet Visit\n7. Preventive Care Tips",
     "Under Signs: Refusal to eat (beyond normal), wheezing, mouth gaping, retained shed, mites, swelling, abnormal stool, stargazing, scale discoloration, lethargy\nUnder Illnesses: RI, IBD, parasites, scale rot, mouth rot\nUnder Preventive: Humidity, temperature, substrate, quarantine new snakes",
     "services.html\nvet-alhambra.html\nBlog #1 (beardie signs)\nBlog #25 (first-time reptile owner)\npricing.html",
     "After Section 5: 'Snake showing respiratory symptoms? Book a reptile exam.'\nEnd: Full CTA",
     "Ball pythons commonly go off feed for months (normal). Distinguish normal from abnormal. Common species in SGV: ball pythons, corn snakes, king snakes, boas."],
]

for row_idx, data in enumerate(content_plans, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx == 1:
            cell.alignment = center_wrap
            cell.font = bold
    ws3.row_dimensions[row_idx].height = 180

# ═══════════════════════════════════════════════════
# SHEET 4: SOCIAL CONTENT FROM BLOG RSS
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Social Posts")

headers4 = ["Blog #", "Blog Title", "Instagram Caption 1", "Instagram Caption 2", "Hashtags", "Visual Concept", "Status"]
widths4 = [8, 35, 55, 55, 45, 40, 14]
style_headers(ws4, headers4, widths4)

social_posts = [
    [1, "Signs Your Bearded Dragon Needs to See a Vet",
     "Is your bearded dragon acting different?\n\nNot eating. Lethargic. Black beard that won't go away.\n\nThese could be signs something's wrong. As vets who see bearded dragons regularly, here are 8 warning signs that mean it's time for a vet visit.\n\nNew blog on our website \u2014 link in bio.\n\n\U0001f4cd 3116 W Main St, Alhambra\n\U0001f4de (626) 441-1314",
     "\u26a0\ufe0f Bearded dragon owners, save this post.\n\nThese are the signs your beardie needs a vet:\n\n1\ufe0f\u20e3 Stopped eating (beyond brumation)\n2\ufe0f\u20e3 Swollen or closed eyes\n3\ufe0f\u20e3 Mouth gaping or mucus\n4\ufe0f\u20e3 Rapid weight loss\n5\ufe0f\u20e3 Dark discoloration on belly\n6\ufe0f\u20e3 Wobbly or unable to walk\n\nDon't wait. Early treatment saves lives.\n\nFull guide on our blog \u2192 link in bio",
     "#BeardedDragon #BeardedDragonCare #ReptileVet #ExoticVet #BeardedDragonHealth #ReptileCare #AlhambraVet #SouthPasadenaAnimalHospital #ReptileLovers #BeardedDragonsOfInstagram #LizardLove #ReptileKeeper #ExoticPetCare #SGV #VetTips",
     "Carousel: Slide 1 = bold text 'Is Your Bearded Dragon Sick?' / Slides 2-7 = each warning sign with icon / Last slide = SPAH logo + Book Now CTA",
     "Not Started"],

    [2, "Guinea Pig Sneezing",
     "Your guinea pig sneezed. Should you panic?\n\nUsually, no. A little sneeze from dust or a new scent is totally normal.\n\nBut if your guinea pig is sneezing repeatedly, has discharge from the nose or eyes, or is eating less \u2014 that could be an upper respiratory infection.\n\nURIs in guinea pigs can be serious. Don't wait.\n\nFull blog post \u2192 link in bio\n\n\U0001f4cd SPAH | 3116 W Main St, Alhambra",
     "Guinea pig sneezing? Here's what to watch for \U0001f447\n\n\u2705 Normal: Occasional sneeze, no other symptoms\n\u26a0\ufe0f See a vet if: Frequent sneezing + runny nose, crusty eyes, loss of appetite, wheezing or clicking sounds, hiding more than usual\n\nGuinea pig respiratory infections can turn serious fast. If you notice these signs, don't wait.\n\nWe treat guinea pigs and exotic pets at SPAH.\n\U0001f4de (626) 441-1314",
     "#GuineaPig #GuineaPigCare #GuineaPigHealth #ExoticVet #SmallAnimalVet #GuineaPigsOfInstagram #CavyLove #PocketPets #ExoticPetCare #AlhambraVet #SGV #VetTips #PetHealth #GuineaPigMom #SouthPasadenaAnimalHospital",
     "Single graphic: 'Normal Sneeze vs. Vet Visit' comparison chart in brand colors. Or carousel with checklist format.",
     "Not Started"],

    [3, "Rabbit GI Stasis",
     "\U0001f6a8 Rabbit owners: if your bunny stops eating, this is an emergency.\n\nGI stasis is one of the most common \u2014 and most dangerous \u2014 conditions in pet rabbits. It can be fatal within 24-48 hours if untreated.\n\nKnow the signs:\n\u2022 Not eating\n\u2022 No droppings (or very small ones)\n\u2022 Hunched posture\n\u2022 Grinding teeth\n\u2022 Bloated belly\n\nIf your rabbit is showing ANY of these signs, call a vet immediately.\n\n\U0001f4de (626) 441-1314\nFull guide \u2192 link in bio",
     "What is GI stasis? And why is it so dangerous for rabbits?\n\nGI stasis happens when your rabbit's digestive system slows down or stops completely. Gas builds up. Bacteria overgrow. Toxins enter the bloodstream.\n\nThe #1 cause? Not enough hay in the diet.\n\n80% of your rabbit's diet should be hay. If it's not, your bunny is at risk.\n\nFull prevention guide on our blog \u2192 link in bio",
     "#Rabbit #RabbitCare #BunnyHealth #GIStasis #RabbitVet #BunniesOfInstagram #HouseRabbit #RabbitHealth #ExoticVet #AlhambraVet #PetRabbit #BunnyLove #RabbitOwner #ExoticPetCare #SGV #SouthPasadenaAnimalHospital",
     "Urgent-styled graphic: Red/orange tones. 'GI STASIS: Know the Signs' header. List symptoms. End slide with 'Call immediately: (626) 441-1314'",
     "Not Started"],

    [4, "Bird Feather Plucking",
     "Feather plucking is one of the most common reasons bird owners visit us.\n\nIf your parrot, cockatoo, or African grey is pulling out its own feathers, it could be:\n\n\U0001f3e5 Medical: skin infection, parasites, liver disease, nutritional deficiency\n\U0001f9e0 Behavioral: boredom, stress, hormonal changes, lack of sleep\n\nThe first step is always a vet exam to rule out medical causes.\n\nDr. Navia sees avian patients at SPAH.\n\U0001f4de (626) 441-1314\nBlog \u2192 link in bio",
     "Why is my bird plucking its feathers? \U0001f99c\n\nBefore you assume it's behavioral, rule out medical causes first.\n\nA simple vet exam can check for:\n\u2022 Skin infections\n\u2022 Parasites\n\u2022 Nutritional deficiencies\n\u2022 Liver or kidney issues\n\nMany cases we see have a medical component that, once treated, reduces or stops the plucking.\n\nDon't guess. Get answers.\nFull guide on our blog \u2192 link in bio",
     "#BirdVet #FeatherPlucking #ParrotCare #AvianVet #BirdHealth #ParrotsOfInstagram #Cockatoo #AfricanGrey #BirdLove #ExoticVet #AlhambraVet #AvianMedicine #PetBird #BirdOwner #SGV #SouthPasadenaAnimalHospital",
     "Carousel: Slide 1 = photo of healthy-feathered bird with question text / Slides = Medical vs Behavioral causes / Last slide = 'Book an avian exam' CTA",
     "Not Started"],

    [5, "Exotic Pet First Vet Visit",
     "Just got a new exotic pet? \U0001f98e\U0001f430\U0001f99c\n\nYour first vet visit should happen within the first 1-2 weeks. Here's why:\n\n\u2022 Many exotic pets hide illness really well\n\u2022 Pet store animals often arrive with parasites or infections\n\u2022 Early detection = cheaper treatment\n\u2022 Your vet can set you up with the right diet, habitat, and care schedule\n\nAt SPAH, we treat reptiles, birds, rabbits, guinea pigs, hamsters, and more.\n\nNew exotic patient? Book your first visit \u2192 link in bio",
     "What actually happens at an exotic pet vet visit? \U0001f447\n\nFull physical exam (head to tail/beak to claw)\nWeight check (exotic pets lose weight fast when sick)\nFecal test for parasites\nDiet & husbandry review\nVaccinations if applicable (rabbits: RHDV2)\n\nNo scary surprises. Just good preventive care.\n\nSee our transparent pricing \u2192 link in bio\n\U0001f4cd SPAH | 3116 W Main St, Alhambra",
     "#ExoticPets #ExoticVet #FirstVetVisit #NewPetOwner #ReptileOwner #BirdOwner #RabbitOwner #GuineaPigOwner #ExoticPetCare #VetVisit #AlhambraVet #PetHealth #ExoticAnimal #SGV #SouthPasadenaAnimalHospital",
     "Reel: Walk-through style 'What happens at an exotic pet vet visit' - show exam room, equipment, handling. Or carousel with each species + what to expect.",
     "Not Started"],

    [6, "Bearded Dragon Care Guide",
     "Southern California is one of the best places to own a bearded dragon.\n\nThe climate, the natural sunlight, the warm temps \u2014 beardies love it here.\n\nBut there's still a lot that can go wrong without proper husbandry.\n\nWe wrote the most comprehensive bearded dragon care guide from a local vet's perspective. Covers:\n\n\u2022 Enclosure setup\n\u2022 UVB lighting\n\u2022 Diet by age\n\u2022 Common health issues\n\u2022 When to see a vet\n\nFull guide \u2192 link in bio",
     "\U0001f321\ufe0f SoCal bearded dragon tip:\n\nYes, our natural sunlight is amazing for beardies. But you STILL need a proper UVB bulb inside the enclosure.\n\nWindow glass filters out UVB rays. So even if your beardie sits by a sunny window, they're not getting what they need.\n\nMore tips in our complete care guide \u2192 link in bio\n\n\U0001f4cd SPAH | Reptile vet in Alhambra",
     "#BeardedDragon #BeardedDragonCare #ReptileCare #BeardedDragonSetup #ReptileKeeper #SoCalPets #BeardedDragonsOfInstagram #ReptileVet #UVBLighting #ReptileHusbandry #AlhambraVet #ExoticVet #SGV #LizardLove #SouthPasadenaAnimalHospital",
     "Carousel: Beautiful beardie photos (source from owner permission or stock). Each slide = one care tip with clean design. Last slide = SPAH CTA.",
     "Not Started"],

    [7, "Toxic Bird Foods",
     "\u274c NEVER feed these to your bird:\n\n1. Avocado (persin = fatal)\n2. Chocolate (theobromine toxicity)\n3. Caffeine (cardiac issues)\n4. Onions & garlic (blood cell damage)\n5. Fruit seeds & pits (cyanide)\n\n\u2705 Safe alternatives:\nBerries, leafy greens, cooked sweet potato, bell peppers, whole grain pasta\n\nSave this post. Share it with a bird owner. It could save a life.\n\nFull list on our blog \u2192 link in bio",
     "Can my bird eat this? \U0001f447\n\n\U0001f34e Apple \u2192 YES (remove seeds!)\n\U0001f951 Avocado \u2192 NEVER \u274c\n\U0001f952 Carrot \u2192 YES \u2705\n\U0001f36b Chocolate \u2192 NEVER \u274c\n\U0001f347 Grapes \u2192 YES (in moderation) \u2705\n\U0001f9c5 Onion \u2192 NEVER \u274c\n\U0001f345 Tomato \u2192 Small amounts OK \u2705\n\nWhen in doubt, ask your avian vet.\n\nFull safe food list on our blog \u2192 link in bio",
     "#BirdDiet #PetBird #BirdCare #ToxicFoodsPets #ParrotFood #AvianNutrition #BirdSafety #ParrotCare #BirdVet #BirdsOfInstagram #ExoticVet #AlhambraVet #PetSafety #BirdOwner #SGV #SouthPasadenaAnimalHospital",
     "Carousel with bold red X / green check format. One food per slide. High save/share potential. Infographic style.",
     "Not Started"],

    [8, "Sick Snake Checklist",
     "Snakes are masters at hiding illness. By the time symptoms are obvious, they may have been sick for weeks.\n\nHere's what to watch for:\n\n\U0001f534 Wheezing or bubbles from nose/mouth\n\U0001f534 Retained shed (especially over eyes)\n\U0001f534 Mouth gaping or excess saliva\n\U0001f534 Visible mites (tiny black/red dots)\n\U0001f534 Swelling anywhere on the body\n\U0001f534 Refusal to eat for 2+ months (non-brumation)\n\n'Not eating' alone isn't always a concern for snakes. But combined with other signs, it's time for a vet visit.\n\nFull checklist \u2192 link in bio",
     "Ball python not eating? \U0001f40d\n\nDon't panic yet. Ball pythons can go months without eating and be perfectly fine \u2014 especially during winter or breeding season.\n\nBUT call a vet if you also notice:\n\u2022 Weight loss\n\u2022 Wheezing\n\u2022 Mouth gaping\n\u2022 Wrinkled skin\n\u2022 Stuck shed\n\nKnow the difference between normal and not.\n\nFull sick snake checklist \u2192 link in bio\n\U0001f4cd SPAH | Reptile vet in Alhambra",
     "#SnakeCare #BallPython #SnakeVet #ReptileHealth #SnakesOfInstagram #ReptileVet #BallPythonCare #ExoticVet #ReptileKeeper #SnakeOwner #PetSnake #ReptileCare #AlhambraVet #SGV #SouthPasadenaAnimalHospital",
     "Carousel: Dark/moody aesthetic matching snake content. Each slide = one warning sign. Or Reel: 'Vet reacts to common snake owner mistakes'",
     "Not Started"],
]

for row_idx, data in enumerate(social_posts, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx in [1, 7]:
            cell.alignment = center_wrap
        if col_idx == 1:
            cell.font = bold
    ws4.row_dimensions[row_idx].height = 250

dv4 = DataValidation(type="list", formula1='"Not Started,Designed,Scheduled,Posted"', allow_blank=True)
ws4.add_data_validation(dv4)
for row in range(2, len(social_posts) + 2):
    dv4.add(ws4.cell(row=row, column=7))

# ═══════════════════════════════════════════════════
# SHEET 5: CMS RECOMMENDATION
# ═══════════════════════════════════════════════════
ws5 = wb.create_sheet("CMS & Platform Strategy")
ws5.column_dimensions["A"].width = 80

lines = [
    ("CMS & BLOGGING PLATFORM RECOMMENDATION", bold12),
    ("", normal),
    ("CURRENT SETUP", Font(bold=True, size=11, name="Arial", color="1F4E79")),
    ("- Static HTML/CSS website hosted on Cloudflare Pages", normal),
    ("- Auto-deploys from GitHub (master branch)", normal),
    ("- Domain: www.spah.la (DNS on Cloudflare)", normal),
    ("- No CMS, no blog infrastructure currently", normal),
    ("", normal),
    ("RECOMMENDED: GHOST (SELF-HOSTED ON CLOUDFLARE OR DIGITAL OCEAN)", Font(bold=True, size=11, name="Arial", color="1F4E79")),
    ("", normal),
    ("Option A: Ghost Blog on Subdirectory (BEST for SEO)", bold),
    ("- Set up Ghost at blog.spah.la or use Cloudflare Workers to serve it at www.spah.la/blog/", normal),
    ("- Subdirectory (/blog/) is better than subdomain (blog.) for SEO authority", normal),
    ("- Ghost is lightweight, fast, SEO-optimized out of the box", normal),
    ("- Built-in RSS feed (needed for social automation)", normal),
    ("- Built-in email newsletters (grow subscriber list)", normal),
    ("- Cost: $9/month (Ghost Pro) or free (self-hosted on $5/month DigitalOcean droplet)", normal),
    ("", normal),
    ("Option B: Add Blog Pages Directly to Static Site", bold),
    ("- Create /blog/ directory with static HTML blog posts", normal),
    ("- Zero cost, same Cloudflare Pages hosting", normal),
    ("- Downside: No CMS, no RSS, manual HTML editing for every post", normal),
    ("- Downside: No newsletter/subscriber functionality", normal),
    ("- Good for first 5-10 posts, but doesn't scale", normal),
    ("", normal),
    ("Option C: WordPress on Subdomain", bold),
    ("- Set up WordPress at blog.spah.la", normal),
    ("- Most popular CMS, huge plugin ecosystem", normal),
    ("- Downside: Subdomain doesn't pass full SEO authority to main site", normal),
    ("- Downside: Requires maintenance, security updates, hosting ($10-30/month)", normal),
    ("- Downside: Slower than Ghost or static unless heavily optimized", normal),
    ("", normal),
    ("RECOMMENDATION: Start with Option B (static blog pages) for the first 5 posts.", Font(bold=True, size=11, name="Arial", color="CC0000")),
    ("This keeps costs at $0 and lets you validate the content strategy.", normal),
    ("Once you see traffic results, upgrade to Ghost for automation and scaling.", normal),
    ("", normal),
    ("SOCIAL AUTOMATION STACK", Font(bold=True, size=11, name="Arial", color="1F4E79")),
    ("", normal),
    ("1. Blog publishes on www.spah.la/blog/[post-slug].html", normal),
    ("2. When Ghost is added: RSS feed auto-generates at www.spah.la/blog/rss/", normal),
    ("3. Blotato connects to RSS feed and auto-creates social drafts", normal),
    ("4. Review drafts in Blotato dashboard > approve > auto-publishes to Instagram + TikTok", normal),
    ("5. Manual review recommended for first 3 months to maintain brand voice", normal),
    ("", normal),
    ("CONTENT PUBLISHING CADENCE", Font(bold=True, size=11, name="Arial", color="1F4E79")),
    ("", normal),
    ("Month 1-2: 2 blog posts/week (P1 posts, exotic species content)", normal),
    ("Month 3-4: 1 blog post/week + 3 social posts/week", normal),
    ("Month 5+: 1 blog post/week + daily social (mix of blog promos + original content)", normal),
    ("", normal),
    ("Each blog post generates:", normal),
    ("- 2 Instagram posts (carousel or single image)", normal),
    ("- 1 Instagram Reel (short-form video adaptation)", normal),
    ("- 1 TikTok (repurpose Reel)", normal),
    ("- 1 Google Business Profile post (local SEO boost)", normal),
    ("", normal),
    ("EXPECTED RESULTS", Font(bold=True, size=11, name="Arial", color="1F4E79")),
    ("", normal),
    ("Month 1-2: Blog posts indexed, start appearing in long-tail searches", normal),
    ("Month 3-4: First page rankings for low-competition exotic keywords", normal),
    ("Month 6: Organic traffic doubles from current baseline", normal),
    ("Month 12: Dominant position for exotic vet content in SGV", normal),
    ("These are conservative estimates given the near-zero competition in your niche.", normal),
]

for i, (text, font) in enumerate(lines, 1):
    cell = ws5.cell(row=i, column=1, value=text)
    cell.font = font
    if font in [Font(bold=True, size=11, name="Arial", color="1F4E79")]:
        cell.fill = PatternFill("solid", fgColor="E3F2FD")

# ═══════════════════════════════════════════════════
# SHEET 6: SPECIES LANDING PAGES (NEW)
# ═══════════════════════════════════════════════════
ws6 = wb.create_sheet("New Landing Pages Needed")

headers6 = ["Page", "Target Keyword", "Supporting Keywords", "SEO Value", "Difficulty", "Status"]
widths6 = [35, 30, 55, 14, 12, 14]
style_headers(ws6, headers6, widths6)

pages = [
    ["reptile-vet-alhambra.html", "reptile vet Alhambra", "reptile vet near me, bearded dragon vet Alhambra, snake vet SGV, lizard vet Los Angeles, turtle vet Alhambra", "Very High", "Low", "Not Started"],
    ["bird-vet-alhambra.html", "bird vet Alhambra", "avian vet near me, parrot vet Pasadena, cockatiel vet SGV, bird vet Los Angeles, avian medicine", "Very High", "Low", "Not Started"],
    ["rabbit-vet-alhambra.html", "rabbit vet Alhambra", "rabbit vet near me, bunny vet SGV, rabbit vet Pasadena, RHDV vaccine, house rabbit vet", "High", "Low", "Not Started"],
    ["guinea-pig-vet-alhambra.html", "guinea pig vet Alhambra", "guinea pig vet near me, cavy vet SGV, exotic small mammal vet, guinea pig checkup", "High", "Low", "Not Started"],
    ["exotic-vet-alhambra.html", "exotic vet Alhambra", "exotic animal hospital, exotic pet vet near me, exotic vet San Gabriel Valley, unusual pet vet", "Very High", "Low", "Not Started"],
    ["exotic-vet-pasadena.html", "exotic vet Pasadena", "exotic vet near Pasadena, reptile vet Pasadena, bird vet Pasadena, exotic animal hospital Pasadena", "High", "Low", "Not Started"],
    ["exotic-vet-los-angeles.html", "exotic vet Los Angeles", "exotic vet LA, reptile vet Los Angeles, bird vet LA, exotic animal vet LA, exotic vet near me", "High", "Medium", "Not Started"],
    ["chicken-vet-los-angeles.html", "chicken vet Los Angeles", "backyard chicken vet, poultry vet LA, chicken vet near me, backyard flock veterinarian", "Moderate", "Low", "Not Started"],
]

seo_fills6 = {
    "Very High": PatternFill("solid", fgColor="C8E6C9"),
    "High": PatternFill("solid", fgColor="DCEDC8"),
    "Moderate": PatternFill("solid", fgColor="FFF9C4"),
}

for row_idx, data in enumerate(pages, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
        if col_idx in [4, 5, 6]:
            cell.alignment = center_wrap
        if col_idx == 4:
            cell.fill = seo_fills6.get(val, PatternFill())
    ws6.row_dimensions[row_idx].height = 50

r6 = len(pages) + 3
note = "NOTE: These species-specific landing pages are CRITICAL. Right now, SPAH doesn't rank for 'exotic vet Alhambra', 'reptile vet SGV', or 'bird vet Pasadena'. These pages will capture that traffic immediately because no competitor has them."
ws6.cell(row=r6, column=1, value=note).font = Font(name="Arial", size=10, italic=True, color="CC0000")
ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=6)

dv6 = DataValidation(type="list", formula1='"Not Started,In Progress,Published"', allow_blank=True)
ws6.add_data_validation(dv6)
for row in range(2, len(pages) + 2):
    dv6.add(ws6.cell(row=row, column=6))

# ═══════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════
output = "C:/Users/rchia/Documents/SPAH-website/SPAH-SEO-Content-Engine.xlsx"
wb.save(output)
print(f"Created: {output}")
