import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Citation Tracker"

# Column headers
headers = ["Priority", "Platform", "Current Address Listed", "Phone Issue", "URL to Update", "How to Update", "Deadline", "Status", "Notes"]
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Column widths
widths = [10, 22, 30, 18, 45, 30, 14, 14, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Priority fills
p1_fill = PatternFill("solid", fgColor="FCE4EC")  # red-ish
p2_fill = PatternFill("solid", fgColor="FFF3E0")  # orange-ish
p3_fill = PatternFill("solid", fgColor="FFFDE7")  # yellow-ish
p4_fill = PatternFill("solid", fgColor="E8F5E9")  # green-ish
p5_fill = PatternFill("solid", fgColor="E3F2FD")  # blue-ish
priority_fills = {"P1": p1_fill, "P2": p2_fill, "P3": p3_fill, "P4": p4_fill, "P5": p5_fill}

red_font = Font(color="CC0000", bold=True, name="Arial", size=10)
normal_font = Font(name="Arial", size=10)

old_addr = "1911 Fremont Ave, South Pasadena, CA 91030"
new_addr = "3116 W Main St, Alhambra, CA 91801"

listings = [
    # P1 - Critical
    ["P1", "Google Business Profile", old_addr, "Correct", "", "Search 'Google Business Profile' > Edit > Update address", "Apr 1", "Not Started", "MOST IMPORTANT - update on move day"],
    ["P1", "Yelp", old_addr, "Correct", "https://biz.yelp.com", "Log in to Yelp for Business > Edit Business Info", "Apr 1", "Not Started", ""],
    ["P1", "Apple Maps (Apple Business Connect)", old_addr, "Correct", "https://businessconnect.apple.com", "Claim/update via Apple Business Connect", "Apr 1", "Not Started", ""],
    ["P1", "Bing Places", old_addr, "Correct", "https://www.bingplaces.com", "Log in > Edit listing > Update address", "Apr 1", "Not Started", ""],
    # P2 - High
    ["P2", "Facebook Page", old_addr, "Correct", "", "Page Settings > Edit Page Info > Address", "Apr 3", "Not Started", ""],
    ["P2", "Instagram Business", old_addr, "Correct", "", "Edit Profile > Business Address", "Apr 3", "Not Started", ""],
    ["P2", "Nextdoor Business Page", old_addr, "Correct", "", "Business dashboard > Edit profile", "Apr 3", "Not Started", ""],
    ["P2", "Waze", old_addr, "Correct", "", "Update via Google Business Profile (syncs)", "Apr 3", "Not Started", "Usually syncs from GBP"],
    # P3 - Medium
    ["P3", "Yellow Pages (YP.com)", old_addr, "WRONG: (626) 441-8002", "https://www.yellowpages.com", "Claim listing > Update address + phone", "Apr 7", "Not Started", "Phone number is WRONG - fix ASAP"],
    ["P3", "Superpages", old_addr, "WRONG: (626) 441-8002", "https://www.superpages.com", "Claim listing > Update (syncs with YP)", "Apr 7", "Not Started", "Phone number is WRONG"],
    ["P3", "MapQuest", old_addr, "Correct", "https://www.mapquest.com", "Submit update request", "Apr 7", "Not Started", ""],
    ["P3", "Foursquare", old_addr, "Correct", "https://foursquare.com", "Claim venue > Edit details", "Apr 7", "Not Started", "Feeds data to Uber, Twitter, etc."],
    ["P3", "Pawlicy Advisor", old_addr, "WRONG: Lists wrong #", "https://www.pawlicy.com", "Contact support to update listing", "Apr 7", "Not Started", "Phone number is WRONG"],
    ["P3", "PetDesk", old_addr, "Correct", "", "Contact PetDesk support or update in app", "Apr 7", "Not Started", ""],
    # P4 - Lower
    ["P4", "Healthgrades (for vets)", old_addr, "Correct", "", "Claim profile > Update", "Apr 14", "Not Started", ""],
    ["P4", "Vitals.com", old_addr, "Correct", "", "Claim profile > Update", "Apr 14", "Not Started", ""],
    ["P4", "Manta", old_addr, "Correct", "https://www.manta.com", "Claim listing > Update", "Apr 14", "Not Started", ""],
    ["P4", "Hotfrog", old_addr, "Correct", "https://www.hotfrog.com", "Add/claim business > Update", "Apr 14", "Not Started", ""],
    ["P4", "Chamberofcommerce.com", old_addr, "Correct", "", "Submit update request", "Apr 14", "Not Started", ""],
    ["P4", "BBB (Better Business Bureau)", old_addr, "Correct", "https://www.bbb.org", "Contact BBB to update listing", "Apr 14", "Not Started", ""],
    # P5 - Data aggregators
    ["P5", "Data Axle (Infogroup)", old_addr, "Correct", "https://www.dataaxle.com", "Submit update - feeds 100+ directories", "Apr 7", "Not Started", "UPDATE EARLY - cascades to many sites"],
    ["P5", "Neustar/Localeze", old_addr, "Correct", "https://www.neustarlocaleze.biz", "Submit update - major data aggregator", "Apr 7", "Not Started", "UPDATE EARLY - cascades"],
    ["P5", "Factual (Foursquare)", old_addr, "Correct", "https://foursquare.com", "Same as Foursquare - submit correction", "Apr 7", "Not Started", "Merged with Foursquare"],
    ["P5", "Acxiom", old_addr, "Correct", "", "Submit via their portal", "Apr 14", "Not Started", ""],
    ["P5", "Angie" + "'" + "s List", old_addr, "Correct", "https://www.angi.com", "Claim/update business profile", "Apr 14", "Not Started", "Now called Angi"],
]

for row_idx, data in enumerate(listings, 2):
    priority = data[0]
    fill = priority_fills.get(priority, PatternFill())
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col_idx <= 2:
            cell.fill = fill
        # Red font for wrong phone numbers
        if col_idx == 4 and "WRONG" in str(val):
            cell.font = red_font

# Summary section
summary_row = len(listings) + 3
ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12, name="Arial")

labels = ["Total Listings:", "Completed:", "In Progress:", "Not Started:", "Completion %:"]
formulas = [
    f'=COUNTA(H2:H{len(listings)+1})',
    f'=COUNTIF(H2:H{len(listings)+1},"Completed")',
    f'=COUNTIF(H2:H{len(listings)+1},"In Progress")',
    f'=COUNTIF(H2:H{len(listings)+1},"Not Started")',
    f'=IF(B{summary_row+1}=0,0,B{summary_row+2}/B{summary_row+1})',
]

for i, (label, formula) in enumerate(zip(labels, formulas)):
    r = summary_row + 1 + i
    ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
    cell = ws.cell(row=r, column=2, value=formula)
    cell.font = Font(name="Arial", size=10)
    if "%" in label:
        cell.number_format = "0%"

# Freeze top row and add filters
ws.auto_filter.ref = f"A1:I{len(listings)+1}"
ws.freeze_panes = "A2"

# Data validation for Status column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,N/A"', allow_blank=True)
dv.error = "Please select a valid status"
dv.errorTitle = "Invalid Status"
ws.add_data_validation(dv)
for row in range(2, len(listings) + 2):
    dv.add(ws.cell(row=row, column=8))

output = "C:/Users/rchia/Documents/SPAH-website/SPAH-Citation-Update-Tracker.xlsx"
wb.save(output)
print(f"Created: {output}")
